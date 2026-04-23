"""
pipeline_to_excel.py
====================
Reads a step_final.json produced by sap_gap_analyser.py and writes a
professional, multi-sheet Excel workbook.

Sheet layout
------------
  1. Master View          – left-join of ALL requirement-linked data (R1 → 84)
  2. Gap Analysis (RICEFW)– gap items only, one row per solution bullet
  3. No-Gap Confirmations – standard-SAP items only
  4. Open Actions         – unrelated to requirements; separate table
  5. Summary              – meeting metadata + counts dashboard

Usage
-----
  python pipeline_to_excel.py --json step_final.json --output SAP_Gap_Analysis.xlsx

  Or import and call:
      from pipeline_to_excel import build_excel
      build_excel("step_final.json", "SAP_Gap_Analysis.xlsx")
"""

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
C_NAVY       = "1F3864"   # header background
C_WHITE      = "FFFFFF"
C_ACCENT     = "2E75B6"   # sub-header / accent
C_GAP_ROW    = "FCE4D6"   # light salmon  – gap rows in Master
C_NOGAP_ROW  = "E2EFDA"   # light green   – no-gap rows in Master
C_ALT_ROW    = "F2F2F2"   # alternating row tint
C_GAP_HEAD   = "C00000"   # red tab / header for gap sheet
C_NOGAP_HEAD = "375623"   # dark green for no-gap sheet
C_ACTION_HEAD= "7030A0"   # purple for open actions
C_SUMMARY_HEAD="2F5496"   # blue for summary

FONT_NAME    = "Arial"
FONT_SIZE    = 10
FONT_SIZE_HD = 11


# ── Style helpers ─────────────────────────────────────────────────────────────
def _hdr(cell, bg=C_NAVY, fg=C_WHITE, bold=True, wrap=True, size=FONT_SIZE_HD):
    cell.font      = Font(name=FONT_NAME, bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=wrap)

def _body(cell, bold=False, wrap=True, align="left", color=None):
    kw = dict(name=FONT_NAME, size=FONT_SIZE, bold=bold)
    if color:
        kw["color"] = color
    cell.font      = Font(**kw)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)

def _fill(cell, hex_color):
    cell.fill = PatternFill("solid", fgColor=hex_color)

def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_border(ws, min_row, max_row, min_col, max_col):
    b = _thin_border()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = b

def _freeze(ws, cell="B2"):
    ws.freeze_panes = cell

def _autofilter(ws, first_row, last_col):
    ws.auto_filter.ref = f"A{first_row}:{get_column_letter(last_col)}{first_row}"

def _col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def _row_height(ws, row_num, h=15):
    ws.row_dimensions[row_num].height = h


# ── Sheet 1: Master View ──────────────────────────────────────────────────────
def _sheet_master(wb, data):
    ws = wb.create_sheet("Master View", 0)
    ws.sheet_view.showGridLines = False

    req_map    = {r["id"]: r for r in data.get("requirements", [])}
    norm_map   = {n["id"]: n for n in data.get("normalized", [])}
    cap_map    = {c["id"]: c for c in data.get("capability_assessment", [])}
    gap_map    = {g["req_id"]: g for g in data.get("gap_analysis", [])}
    nogap_map  = {n["id"]: n for n in data.get("no_gap_confirmations", [])}

    # ── Group headers (row 1) ──────────────────────────────────────────────
    GROUP_HEADERS = [
        # (label, start_col, span, bg)
        ("Step 1 — Requirement",           1, 2, C_NAVY),
        ("Step 2 — Normalization",          3, 4, C_ACCENT),
        ("Steps 3 & 4 — SAP Assessment",    7, 3, "4472C4"),
        ("Step 5-7 — RICEFW / Gap Detail", 10, 5, C_GAP_HEAD),
        ("No-Gap Confirmation",            15, 2, C_NOGAP_HEAD),
    ]
    for label, start, span, bg in GROUP_HEADERS:
        ws.merge_cells(
            start_row=1, start_column=start,
            end_row=1,   end_column=start + span - 1
        )
        c = ws.cell(row=1, column=start, value=label)
        _hdr(c, bg=bg, size=FONT_SIZE_HD)
        _row_height(ws, 1, 20)

    # ── Column headers (row 2) ─────────────────────────────────────────────
    COL_HEADERS = [
        "Req ID", "Requirement Text",                      # Step 1
        "Actor", "Action", "Object", "Condition",          # Step 2
        "SAP Status", "Gap?", "Assessment Note",           # Steps 3&4
        "GAP ID", "GAP Title", "RICEFW",                   # Step 5-7 (gap)
        "Solution Strategy",
        "SAP Mechanism(s)",
        "No-Gap Topic", "Standard SAP Resolution",         # No-gap
    ]
    for col_idx, label in enumerate(COL_HEADERS, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        # Match bg to group
        if   col_idx <= 2:  bg = C_NAVY
        elif col_idx <= 6:  bg = C_ACCENT
        elif col_idx <= 9:  bg = "4472C4"
        elif col_idx <= 14: bg = C_GAP_HEAD
        else:               bg = C_NOGAP_HEAD
        _hdr(c, bg=bg, size=FONT_SIZE, wrap=True)
        _row_height(ws, 2, 32)

    _autofilter(ws, 2, len(COL_HEADERS))
    _freeze(ws, "C3")

    # ── Data rows ──────────────────────────────────────────────────────────
    all_ids = sorted(req_map.keys(), key=lambda x: int(x[1:]))
    for row_idx, rid in enumerate(all_ids, start=3):
        req   = req_map.get(rid, {})
        norm  = norm_map.get(rid, {})
        cap   = cap_map.get(rid, {})
        gap   = gap_map.get(rid, {})
        nogap = nogap_map.get(rid, {})

        is_gap  = bool(cap.get("gap"))
        status  = cap.get("status", "")
        sol_txt = "\n".join(
            f"{i}. {b}" for i, b in enumerate(gap.get("solution_bullets", []), 1)
        )
        mechs = _extract_mechs(gap.get("solution_bullets", []))

        row_data = [
            rid,
            req.get("text", ""),
            norm.get("actor", ""),
            norm.get("action", ""),
            norm.get("object", ""),
            norm.get("condition", ""),
            status,
            "YES" if is_gap else "NO",
            cap.get("assessment_note", ""),
            gap.get("gap_id", ""),
            gap.get("title", ""),
            gap.get("ricefw", ""),
            sol_txt,
            mechs,
            nogap.get("topic", ""),
            nogap.get("resolution", ""),
        ]

        row_bg = C_GAP_ROW if is_gap else C_NOGAP_ROW
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            _body(c)
            _fill(c, row_bg)
            if col_idx == 1:
                _body(c, bold=True, align="center")
            if col_idx == 7:
                color = C_GAP_HEAD if status == "NONE" else (
                    "E36209" if status == "PARTIAL" else C_NOGAP_HEAD
                )
                c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=color)
            if col_idx == 8:
                c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True,
                              color=C_GAP_HEAD if is_gap else C_NOGAP_HEAD)

        # Alternate banding for un-tinted cols
        if row_idx % 2 == 0:
            pass  # gap/nogap fill is more informative; skip band

    _apply_border(ws, 1, ws.max_row, 1, len(COL_HEADERS))

    # ── Column widths ──────────────────────────────────────────────────────
    _col_widths(ws, {
        "A": 8,  "B": 42, "C": 12, "D": 14, "E": 28, "F": 22,
        "G": 10, "H": 7,  "I": 38, "J": 10, "K": 30, "L": 14,
        "M": 50, "N": 30, "O": 22, "P": 38,
    })

    # Row heights for data rows
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    # Legend box
    row = ws.max_row + 2
    ws.cell(row=row,   column=1, value="LEGEND").font = Font(bold=True, name=FONT_NAME)
    for col, label, bg in [
        (2, "GAP — PARTIAL (needs enhancement)", C_GAP_ROW),
        (3, "GAP — NONE (custom development)",   "F4B8A8"),
        (4, "NO GAP — Standard SAP",             C_NOGAP_ROW),
    ]:
        c = ws.cell(row=row, column=col, value=label)
        _fill(c, bg)
        c.font = Font(name=FONT_NAME, size=9)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width, 30
        )


# ── Sheet 2: Gap Analysis (RICEFW) ────────────────────────────────────────────
def _sheet_gaps(wb, data):
    ws = wb.create_sheet("Gap Analysis (RICEFW)")
    ws.sheet_view.showGridLines = False

    req_map = {r["id"]: r["text"] for r in data.get("requirements", [])}
    cap_map = {c["id"]: c for c in data.get("capability_assessment", [])}
    gaps    = data.get("gap_analysis", [])

    COLS = ["GAP ID", "Req ID", "Requirement Text",
            "SAP Status", "Assessment Note",
            "RICEFW Type", "GAP Title",
            "#", "Solution Bullet"]

    # Title
    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1,
                value=f"Gap Analysis & RICEFW Classification  —  {len(gaps)} gaps identified")
    _hdr(c, bg=C_GAP_HEAD, size=12)
    ws.row_dimensions[1].height = 24

    for col_idx, label in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        _hdr(c, bg=C_GAP_HEAD, size=FONT_SIZE)
    ws.row_dimensions[2].height = 28

    _autofilter(ws, 2, len(COLS))
    _freeze(ws, "C3")

    row_idx = 3
    for g in gaps:
        rid    = g.get("req_id", "")
        req_t  = req_map.get(rid, "")
        cap    = cap_map.get(rid, {})
        status = cap.get("status", "")
        bullets = g.get("solution_bullets", [])
        span   = max(len(bullets), 1)
        bg     = "F4B8A8" if status == "NONE" else C_GAP_ROW

        shared = [
            g.get("gap_id", ""), rid, req_t,
            status, cap.get("assessment_note", ""),
            g.get("ricefw", ""), g.get("title", ""),
        ]

        for b_idx, bullet in enumerate(bullets or [""]):
            for col_idx, val in enumerate(shared, start=1):
                c = ws.cell(row=row_idx + b_idx, column=col_idx,
                            value=val if b_idx == 0 else "")
                _body(c, wrap=True)
                _fill(c, bg)
                if col_idx in (1, 2, 6) and b_idx == 0:
                    _body(c, bold=True, align="center")

            ws.cell(row=row_idx + b_idx, column=8, value=b_idx + 1 if bullet else "")
            bc = ws.cell(row=row_idx + b_idx, column=9, value=bullet)
            _body(bc, wrap=True)
            _fill(bc, bg)

        # Merge shared columns across bullet rows
        if span > 1:
            for col_idx in range(1, 8):
                ws.merge_cells(
                    start_row=row_idx, start_column=col_idx,
                    end_row=row_idx + span - 1, end_column=col_idx
                )
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    vertical="center", wrap_text=True
                )

        row_idx += span

    _apply_border(ws, 1, ws.max_row, 1, len(COLS))

    _col_widths(ws, {
        "A": 10, "B": 8,  "C": 42, "D": 10,
        "E": 40, "F": 14, "G": 30, "H": 4, "I": 60,
    })
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 40


# ── Sheet 3: No-Gap Confirmations ─────────────────────────────────────────────
def _sheet_nogap(wb, data):
    ws = wb.create_sheet("No-Gap Confirmations")
    ws.sheet_view.showGridLines = False

    req_map  = {r["id"]: r["text"] for r in data.get("requirements", [])}
    norm_map = {n["id"]: n for n in data.get("normalized", [])}
    cap_map  = {c["id"]: c for c in data.get("capability_assessment", [])}
    nogaps   = data.get("no_gap_confirmations", [])

    COLS = ["Req ID", "Topic", "Requirement Text",
            "Actor", "Action",
            "SAP Assessment Note", "Standard SAP Resolution"]

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1,
                value=f"Standard SAP Coverage  —  {len(nogaps)} requirements fully covered")
    _hdr(c, bg=C_NOGAP_HEAD, size=12)
    ws.row_dimensions[1].height = 24

    for col_idx, label in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        _hdr(c, bg=C_NOGAP_HEAD, size=FONT_SIZE)
    ws.row_dimensions[2].height = 28

    _autofilter(ws, 2, len(COLS))
    _freeze(ws, "B3")

    for row_idx, ng in enumerate(nogaps, start=3):
        rid  = ng.get("id", "")
        norm = norm_map.get(rid, {})
        cap  = cap_map.get(rid, {})
        bg   = C_NOGAP_ROW if row_idx % 2 == 0 else "F0FFF0"

        row_data = [
            rid,
            ng.get("topic", ""),
            req_map.get(rid, ""),
            norm.get("actor", ""),
            norm.get("action", ""),
            cap.get("assessment_note", ""),
            ng.get("resolution", ""),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            _body(c, wrap=True)
            _fill(c, bg)
            if col_idx == 1:
                _body(c, bold=True, align="center")

    _apply_border(ws, 1, ws.max_row, 1, len(COLS))

    _col_widths(ws, {
        "A": 8, "B": 25, "C": 45, "D": 12,
        "E": 16, "F": 42, "G": 45,
    })
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 45


# ── Sheet 4: Open Actions ─────────────────────────────────────────────────────
def _sheet_actions(wb, data):
    ws = wb.create_sheet("Open Actions")
    ws.sheet_view.showGridLines = False

    actions = data.get("open_actions", [])

    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value="Open Actions & Next Steps")
    _hdr(c, bg=C_ACTION_HEAD, size=12)
    ws.row_dimensions[1].height = 24

    COLS = ["#", "Action Description", "Owner / Responsible Team",
            "Target / Timing", "Status"]

    for col_idx, label in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        _hdr(c, bg=C_ACTION_HEAD, size=FONT_SIZE)
    ws.row_dimensions[2].height = 28

    _freeze(ws, "B3")

    for row_idx, action in enumerate(actions, start=3):
        bg = "F3E8FF" if row_idx % 2 == 0 else "EDD9FF"
        row_data = [
            action.get("action_number", ""),
            action.get("description", ""),
            action.get("owner", ""),
            action.get("target", ""),
            "Open",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            _body(c, wrap=True)
            _fill(c, bg)
            if col_idx == 1:
                _body(c, bold=True, align="center")
            if col_idx == 5:
                c.font = Font(name=FONT_NAME, size=FONT_SIZE,
                              bold=True, color=C_ACTION_HEAD)

    _apply_border(ws, 1, ws.max_row, 1, len(COLS))
    _col_widths(ws, {
        "A": 5, "B": 65, "C": 30, "D": 18, "E": 12
    })
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 50


# ── Sheet 5: Summary Dashboard ────────────────────────────────────────────────
def _sheet_summary(wb, data):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    reqs     = data.get("requirements", [])
    caps     = data.get("capability_assessment", [])
    gaps     = data.get("gap_analysis", [])
    nogaps   = data.get("no_gap_confirmations", [])
    actions  = data.get("open_actions", [])

    total    = len(reqs)
    n_gap    = sum(1 for c in caps if c.get("gap"))
    n_full   = sum(1 for c in caps if c.get("status") == "FULL")
    n_partial= sum(1 for c in caps if c.get("status") == "PARTIAL")
    n_none   = sum(1 for c in caps if c.get("status") == "NONE")

    ricefw_counts = {}
    for g in gaps:
        rf = g.get("ricefw", "Unknown")
        ricefw_counts[rf] = ricefw_counts.get(rf, 0) + 1

    # Title block
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="SAP IS-U Gap Analysis — Executive Summary")
    c.font      = Font(name=FONT_NAME, bold=True, size=16, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=C_SUMMARY_HEAD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1,
            value=f"{data.get('meeting_title','')}  |  {data.get('meeting_date','')}").font = Font(
        name=FONT_NAME, size=11, italic=True, color=C_SUMMARY_HEAD)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:F3")
    scope_c = ws.cell(row=3, column=1, value=data.get("scope_context", ""))
    scope_c.font      = Font(name=FONT_NAME, size=9, italic=True)
    scope_c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[3].height = 55

    # KPI block ─────────────────────────────────────────────────────────────
    def _kpi_block(row, col, label, value, bg, fg=C_WHITE):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row+1, end_column=col)
        vc = ws.cell(row=row, column=col, value=value)
        vc.font      = Font(name=FONT_NAME, bold=True, size=22, color=fg)
        vc.fill      = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")

        lc = ws.cell(row=row+2, column=col, value=label)
        lc.font      = Font(name=FONT_NAME, size=9, bold=True, color=bg)
        lc.alignment = Alignment(horizontal="center")

    KPI_ROW = 5
    _kpi_block(KPI_ROW, 1, "Total Requirements", total,       C_SUMMARY_HEAD)
    _kpi_block(KPI_ROW, 2, "Gaps Identified",    n_gap,       C_GAP_HEAD)
    _kpi_block(KPI_ROW, 3, "Standard SAP (FULL)",n_full,      C_NOGAP_HEAD)
    _kpi_block(KPI_ROW, 4, "Partial Gaps",        n_partial,  "E36209")
    _kpi_block(KPI_ROW, 5, "Custom Dev Required", n_none,     "7B0000")
    _kpi_block(KPI_ROW, 6, "Open Actions",        len(actions),C_ACTION_HEAD)

    for r in range(KPI_ROW, KPI_ROW + 3):
        ws.row_dimensions[r].height = 24

    # RICEFW breakdown ─────────────────────────────────────────────────────
    row = KPI_ROW + 4
    ws.cell(row=row, column=1, value="RICEFW Breakdown").font = Font(
        name=FONT_NAME, bold=True, size=11, color=C_SUMMARY_HEAD)
    ws.row_dimensions[row].height = 20
    row += 1

    ricefw_colors = {
        "Report":      "9DC3E6", "Interface":   "F4B942",
        "Conversion":  "A9D18E", "Enhancement": "F4B8A8",
        "Form":        "FFE699", "Workflow":    "D5A6BD",
    }
    for col_idx, (rf, cnt) in enumerate(sorted(ricefw_counts.items()), start=1):
        hc = ws.cell(row=row,   column=col_idx, value=rf)
        vc = ws.cell(row=row+1, column=col_idx, value=cnt)
        bg = ricefw_colors.get(rf, "E0E0E0")
        _hdr(hc, bg=C_NAVY, size=9)
        _fill(vc, bg)
        vc.font      = Font(name=FONT_NAME, bold=True, size=14)
        vc.alignment = Alignment(horizontal="center")

    ws.row_dimensions[row].height   = 18
    ws.row_dimensions[row+1].height = 24

    # Gap list quick-ref ───────────────────────────────────────────────────
    row = row + 3
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="Gap Quick Reference").font = Font(
        name=FONT_NAME, bold=True, size=11, color=C_SUMMARY_HEAD)
    ws.row_dimensions[row].height = 20
    row += 1

    for col_idx, label in enumerate(
        ["GAP ID", "Req ID", "Title", "RICEFW", "SAP Status", "Sheet Ref"], start=1
    ):
        c = ws.cell(row=row, column=col_idx, value=label)
        _hdr(c, bg=C_GAP_HEAD, size=FONT_SIZE)
    ws.row_dimensions[row].height = 20
    row += 1

    cap_map = {c["id"]: c for c in caps}
    for g in data.get("gap_analysis", []):
        cap   = cap_map.get(g.get("req_id", ""), {})
        bg    = "F4B8A8" if cap.get("status") == "NONE" else C_GAP_ROW
        cells = [g.get("gap_id",""), g.get("req_id",""), g.get("title",""),
                 g.get("ricefw",""), cap.get("status",""), "Gap Analysis (RICEFW)"]
        for col_idx, val in enumerate(cells, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            _body(c)
            _fill(c, bg)
        row += 1

    _apply_border(ws, 4, ws.max_row, 1, 6)

    _col_widths(ws, {
        "A": 12, "B": 10, "C": 40, "D": 16, "E": 14, "F": 24
    })


# ── SAP mechanism extractor ───────────────────────────────────────────────────
def _extract_mechs(bullets: list) -> str:
    """Pull SAP object names (ALL-CAPS words / known patterns) from bullets."""
    import re
    keywords = set()
    patterns = [
        r"\bBAPI_\w+",
        r"\bBAdI\s+\w+",
        r"\bBRF\+",
        r"\bCDS\b",
        r"\bOData\b",
        r"\bIDOC\s+\w+",
        r"\bIDoc\s+\w+",
        r"\bLSMW\b",
        r"\bFICA?\b",
        r"\bFPE1\b",
        r"\bFPG1\b",
        r"\bFQZ0\b",
        r"\bEASABI\b",
        r"\bEASIBI\b",
        r"\bEMIGALL\b",
        r"\bABAP\b",
        r"\bEC50E\b",
        r"\bFiori\b",
        r"\bPFCG\b",
        r"\bEMMA\b",
        r"\bEDM\b",
        r"\bRBAC\b",
        r"ISU_\w+",
        r"FKK_\w+",
        r"CL_\w+",
    ]
    for bullet in bullets:
        for pat in patterns:
            keywords.update(re.findall(pat, bullet, re.IGNORECASE))
    return ", ".join(sorted(keywords)) if keywords else ""


# ── Main builder ─────────────────────────────────────────────────────────────
def build_excel(json_path: str, output_path: str) -> str:
    data = json.loads(Path(json_path).read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    _sheet_summary(wb, data)
    _sheet_master(wb, data)
    _sheet_gaps(wb, data)
    _sheet_nogap(wb, data)
    _sheet_actions(wb, data)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"[OK] Written → {output_path}")
    return output_path


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert sap_gap_analyser step_final.json to Excel"
    )
    parser.add_argument("--json",   required=True, help="Path to step_final.json")
    parser.add_argument("--output", default="SAP_Gap_Analysis.xlsx",
                        help="Output .xlsx path")
    args = parser.parse_args()
    build_excel(args.json, args.output)
