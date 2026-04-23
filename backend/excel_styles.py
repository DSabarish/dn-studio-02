from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

C_NAVY = "1F3864"
C_WHITE = "FFFFFF"
C_ACCENT = "2E75B6"
C_GAP_ROW = "FCE4D6"
C_NOGAP_ROW = "E2EFDA"
C_GAP_HEAD = "C00000"
C_NOGAP_HEAD = "375623"
C_ACTION_HEAD = "7030A0"
C_SUMMARY_HEAD = "2F5496"

FONT_NAME = "Arial"
FONT_SIZE = 10
FONT_SIZE_HD = 11


def hdr(cell, bg=C_NAVY, fg=C_WHITE, bold=True, wrap=True, size=FONT_SIZE_HD):
    cell.font = Font(name=FONT_NAME, bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def body(cell, bold=False, wrap=True, align="left", color=None):
    kw = dict(name=FONT_NAME, size=FONT_SIZE, bold=bold)
    if color:
        kw["color"] = color
    cell.font = Font(**kw)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)


def fill(cell, hex_color):
    cell.fill = PatternFill("solid", fgColor=hex_color)


def thin_border() -> Border:
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def apply_border(ws, min_row, max_row, min_col, max_col):
    border = thin_border()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def freeze(ws, cell="B2"):
    ws.freeze_panes = cell


def autofilter(ws, first_row, last_col):
    ws.auto_filter.ref = f"A{first_row}:{get_column_letter(last_col)}{first_row}"


def col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def row_height(ws, row_num, height=15):
    ws.row_dimensions[row_num].height = height
