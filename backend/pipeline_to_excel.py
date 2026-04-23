"""
pipeline_to_excel.py
====================
Reads a step_final.json produced by sap_gap_analyser.py and writes a
professional, multi-sheet Excel workbook.

Sheet layout
------------
  1. Master View          – left-join of ALL requirement-linked data (R1 → 84)
  2. Gap Analysis (RICEFW)– gap items only, one row per solution bullet
  3. No-Gap Confirmations – standard-SAP items only
  4. Open Actions         – unrelated to requirements; separate table
  5. Summary              – meeting metadata + counts dashboard

Usage
-----
  python pipeline_to_excel.py --json step_final.json --output SAP_Gap_Analysis.xlsx

  Or import and call:
      from pipeline_to_excel import build_excel
      build_excel("step_final.json", "SAP_Gap_Analysis.xlsx")
"""

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from backend.excel_styles import (
    C_ACCENT,
    C_ACTION_HEAD,
    C_GAP_HEAD,
    C_GAP_ROW,
    C_NAVY,
    C_NOGAP_HEAD,
    C_NOGAP_ROW,
    C_SUMMARY_HEAD,
    C_WHITE,
    FONT_NAME,
    FONT_SIZE,
    FONT_SIZE_HD,
    apply_border,
    autofilter,
    body,
    col_widths,
    fill,
    freeze,
    hdr,
    row_height,
)
from backend.excel_utils import extract_mechanisms


# ── Sheet 1: Master View ──────────────────────────────────────────────────────
def _sheet_master(wb, data):
    ws = wb.create_sheet("Master View", 0)
    ws.sheet_view.showGridLines = False

    req_map    = {r["id"]: r for r in data.get("requirements", [])}
    norm_map   = {n["id"]: n for n in data.get("normalized", [])}
    cap_map    = {c["id"]: c for c in data.get("capability_assessment", [])}
    gap_map    = {g["req_id"]: g for g in data.get("gap_analysis", [])}
    nogap_map  = {n["id"]: n for n in data.get("no_gap_confirmations", [])}

    # ── Group headers (row 1) ──────────────────────────────────────────────
    GROUP_HEADERS = [
        # (label, start_col, span, bg)
        ("Step 1 — Requirement",           1, 2, C_NAVY),
        ("Step 2 — Normalization",          3, 4, C_ACCENT),
        ("Steps 3 & 4 — SAP Assessment",    7, 3, "4472C4"),
        ("Step 5-7 — RICEFW / Gap Detail", 10, 5, C_GAP_HEAD),
        ("No-Gap Confirmation",            15, 2, C_NOGAP_HEAD),
    ]
    for label, start, span, bg in GROUP_HEADERS:
        ws.merge_cells(
            start_row=1, start_column=start,
            end_row=1,   end_column=start + span - 1
        )
        c = ws.cell(row=1, column=start, value=label)
        hdr(c, bg=bg, size=FONT_SIZE_HD)
        row_height(ws, 1, 20)

    # ── Column headers (row 2) ─────────────────────────────────────────────
    COL_HEADERS = [
        "Req ID", "Requirement Text",                      # Step 1
        "Actor", "Action", "Object", "Condition",          # Step 2
        "SAP Status", "Gap?", "Assessment Note",           # Steps 3&4
        "GAP ID", "GAP Title", "RICEFW",                   # Step 5-7 (gap)
        "Solution Strategy",
        "SAP Mechanism(s)",
        "No-Gap Topic", "Standard SAP Resolution",         # No-gap
    ]
    for col_idx, label in enumerate(COL_HEADERS, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        # Match bg to group
        if   col_idx <= 2:  bg = C_NAVY
        elif col_idx <= 6:  bg = C_ACCENT
        elif col_idx <= 9:  bg = "4472C4"
        elif col_idx <= 14: bg = C_GAP_HEAD
        else:               bg = C_NOGAP_HEAD
        hdr(c, bg=bg, size=FONT_SIZE, wrap=True)
        row_height(ws, 2, 32)

    autofilter(ws, 2, len(COL_HEADERS))
    freeze(ws, "C3")

    # ── Data rows ──────────────────────────────────────────────────────────
    all_ids = sorted(req_map.keys(), key=lambda x: int(x[1:]))
    for row_idx, rid in enumerate(all_ids, start=3):
        req   = req_map.get(rid, {})
        norm  = norm_map.get(rid, {})
        cap   = cap_map.get(rid, {})
        gap   = gap_map.get(rid, {})
        nogap = nogap_map.get(rid, {})

        is_gap  = bool(cap.get("gap"))
        status  = cap.get("status", "")
        sol_txt = "\n".join(
            f"{i}. {b}" for i, b in enumerate(gap.get("solution_bullets", []), 1)
        )
        mechs = extract_mechanisms(gap.get("solution_bullets", []))

        row_data = [
            rid,
            req.get("text", ""),
            norm.get("actor", ""),
            norm.get("action", ""),
            norm.get("object", ""),
            norm.get("condition", ""),
            status,
            "YES" if is_gap else "NO",
            cap.get("assessment_note", ""),
            gap.get("gap_id", ""),
            gap.get("title", ""),
            gap.get("ricefw", ""),
            sol_txt,
            mechs,
            nogap.get("topic", ""),
            nogap.get("resolution", ""),
        ]

        row_bg = C_GAP_ROW if is_gap else C_NOGAP_ROW
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            body(c)
            fill(c, row_bg)
            if col_idx == 1:
                body(c, bold=True, align="center")
            if col_idx == 7:
                color = C_GAP_HEAD if status == "NONE" else (
                    "E36209" if status == "PARTIAL" else C_NOGAP_HEAD
                )
                c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=color)
            if col_idx == 8:
                c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True,
                              color=C_GAP_HEAD if is_gap else C_NOGAP_HEAD)

        # Alternate banding for un-tinted cols
        if row_idx % 2 == 0:
            pass  # gap/nogap fill is more informative; skip band

    apply_border(ws, 1, ws.max_row, 1, len(COL_HEADERS))

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths(ws, {
        "A": 8,  "B": 42, "C": 12, "D": 14, "E": 28, "F": 22,
        "G": 10, "H": 7,  "I": 38, "J": 10, "K": 30, "L": 14,
        "M": 50, "N": 30, "O": 22, "P": 38,
    })

    # Row heights for data rows
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    # Legend box
    row = ws.max_row + 2
    ws.cell(row=row,   column=1, value="LEGEND").font = Font(bold=True, name=FONT_NAME)
    for col, label, bg in [
        (2, "GAP — PARTIAL (needs enhancement)", C_GAP_ROW),
        (3, "GAP — NONE (custom development)",   "F4B8A8"),
        (4, "NO GAP — Standard SAP",             C_NOGAP_ROW),
    ]:
        c = ws.cell(row=row, column=col, value=label)
        fill(c, bg)
        c.font = Font(name=FONT_NAME, size=9)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width, 30
        )


# ── Sheet 2: Gap Analysis (RICEFW) ────────────────────────────────────────────
def _sheet_gaps(wb, data):
    ws = wb.create_sheet("Gap Analysis (RICEFW)")
    ws.sheet_view.showGridLines = False

    req_map = {r["id"]: r["text"] for r in data.get("requirements", [])}
    cap_map = {c["id"]: c for c in data.get("capability_assessment", [])}
    gaps    = data.get("gap_analysis", [])

    COLS = ["GAP ID", "Req ID", "Requirement Text",
            "SAP Status", "Assessment Note",
            "RICEFW Type", "GAP Title",
            "#", "Solution Bullet"]

    # Title
    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1,
                value=f"Gap Analysis & RICEFW Classification  —  {len(gaps)} gaps identified")
    hdr(c, bg=C_GAP_HEAD, size=12)
    ws.row_dimensions[1].height = 24

    for col_idx, label in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        hdr(c, bg=C_GAP_HEAD, size=FONT_SIZE)
    ws.row_dimensions[2].height = 28

    autofilter(ws, 2, len(COLS))
    freeze(ws, "C3")

    row_idx = 3
    for g in gaps:
        rid    = g.get("req_id", "")
        req_t  = req_map.get(rid, "")
        cap    = cap_map.get(rid, {})
        status = cap.get("status", "")
        bullets = g.get("solution_bullets", [])
        span   = max(len(bullets), 1)
        bg     = "F4B8A8" if status == "NONE" else C_GAP_ROW

        shared = [
            g.get("gap_id", ""), rid, req_t,
            status, cap.get("assessment_note", ""),
            g.get("ricefw", ""), g.get("title", ""),
        ]

        for b_idx, bullet in enumerate(bullets or [""]):
            for col_idx, val in enumerate(shared, start=1):
                c = ws.cell(row=row_idx + b_idx, column=col_idx,
                            value=val if b_idx == 0 else "")
                body(c, wrap=True)
                fill(c, bg)
                if col_idx in (1, 2, 6) and b_idx == 0:
                    body(c, bold=True, align="center")

            ws.cell(row=row_idx + b_idx, column=8, value=b_idx + 1 if bullet else "")
            bc = ws.cell(row=row_idx + b_idx, column=9, value=bullet)
            body(bc, wrap=True)
            fill(bc, bg)

        # Merge shared columns across bullet rows
        if span > 1:
            for col_idx in range(1, 8):
                ws.merge_cells(
                    start_row=row_idx, start_column=col_idx,
                    end_row=row_idx + span - 1, end_column=col_idx
                )
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    vertical="center", wrap_text=True
                )

        row_idx += span

    apply_border(ws, 1, ws.max_row, 1, len(COLS))

    col_widths(ws, {
        "A": 10, "B": 8,  "C": 42, "D": 10,
        "E": 40, "F": 14, "G": 30, "H": 4, "I": 60,
    })
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 40


# ── Sheet 3: No-Gap Confirmations ─────────────────────────────────────────────
def _sheet_nogap(wb, data):
    ws = wb.create_sheet("No-Gap Confirmations")
    ws.sheet_view.showGridLines = False

    req_map  = {r["id"]: r["text"] for r in data.get("requirements", [])}
    norm_map = {n["id"]: n for n in data.get("normalized", [])}
    cap_map  = {c["id"]: c for c in data.get("capability_assessment", [])}
    nogaps   = data.get("no_gap_confirmations", [])

    COLS = ["Req ID", "Topic", "Requirement Text",
            "Actor", "Action",
            "SAP Assessment Note", "Standard SAP Resolution"]

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1,
                value=f"Standard SAP Coverage  —  {len(nogaps)} requirements fully covered")
    hdr(c, bg=C_NOGAP_HEAD, size=12)
    ws.row_dimensions[1].height = 24

    for col_idx, label in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        hdr(c, bg=C_NOGAP_HEAD, size=FONT_SIZE)
    ws.row_dimensions[2].height = 28

    autofilter(ws, 2, len(COLS))
    freeze(ws, "B3")

    for row_idx, ng in enumerate(nogaps, start=3):
        rid  = ng.get("id", "")
        norm = norm_map.get(rid, {})
        cap  = cap_map.get(rid, {})
        bg   = C_NOGAP_ROW if row_idx % 2 == 0 else "F0FFF0"

        row_data = [
            rid,
            ng.get("topic", ""),
            req_map.get(rid, ""),
            norm.get("actor", ""),
            norm.get("action", ""),
            cap.get("assessment_note", ""),
            ng.get("resolution", ""),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            body(c, wrap=True)
            fill(c, bg)
            if col_idx == 1:
                body(c, bold=True, align="center")

    apply_border(ws, 1, ws.max_row, 1, len(COLS))

    col_widths(ws, {
        "A": 8, "B": 25, "C": 45, "D": 12,
        "E": 16, "F": 42, "G": 45,
    })
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 45


# ── Sheet 4: Open Actions ─────────────────────────────────────────────────────
def _sheet_actions(wb, data):
    ws = wb.create_sheet("Open Actions")
    ws.sheet_view.showGridLines = False

    actions = data.get("open_actions", [])

    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value="Open Actions & Next Steps")
    hdr(c, bg=C_ACTION_HEAD, size=12)
    ws.row_dimensions[1].height = 24

    COLS = ["#", "Action Description", "Owner / Responsible Team",
            "Target / Timing", "Status"]

    for col_idx, label in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        hdr(c, bg=C_ACTION_HEAD, size=FONT_SIZE)
    ws.row_dimensions[2].height = 28

    freeze(ws, "B3")

    for row_idx, action in enumerate(actions, start=3):
        bg = "F3E8FF" if row_idx % 2 == 0 else "EDD9FF"
        row_data = [
            action.get("action_number", ""),
            action.get("description", ""),
            action.get("owner", ""),
            action.get("target", ""),
            "Open",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            body(c, wrap=True)
            fill(c, bg)
            if col_idx == 1:
                body(c, bold=True, align="center")
            if col_idx == 5:
                c.font = Font(name=FONT_NAME, size=FONT_SIZE,
                              bold=True, color=C_ACTION_HEAD)

    apply_border(ws, 1, ws.max_row, 1, len(COLS))
    col_widths(ws, {
        "A": 5, "B": 65, "C": 30, "D": 18, "E": 12
    })
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 50


# ── Sheet 5: Summary Dashboard ────────────────────────────────────────────────
def _sheet_summary(wb, data):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    reqs     = data.get("requirements", [])
    caps     = data.get("capability_assessment", [])
    gaps     = data.get("gap_analysis", [])
    nogaps   = data.get("no_gap_confirmations", [])
    actions  = data.get("open_actions", [])

    total    = len(reqs)
    n_gap    = sum(1 for c in caps if c.get("gap"))
    n_full   = sum(1 for c in caps if c.get("status") == "FULL")
    n_partial= sum(1 for c in caps if c.get("status") == "PARTIAL")
    n_none   = sum(1 for c in caps if c.get("status") == "NONE")

    ricefw_counts = {}
    for g in gaps:
        rf = g.get("ricefw", "Unknown")
        ricefw_counts[rf] = ricefw_counts.get(rf, 0) + 1

    # Title block
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="SAP IS-U Gap Analysis — Executive Summary")
    c.font      = Font(name=FONT_NAME, bold=True, size=16, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=C_SUMMARY_HEAD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1,
            value=f"{data.get('meeting_title','')}  |  {data.get('meeting_date','')}").font = Font(
        name=FONT_NAME, size=11, italic=True, color=C_SUMMARY_HEAD)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:F3")
    scope_c = ws.cell(row=3, column=1, value=data.get("scope_context", ""))
    scope_c.font      = Font(name=FONT_NAME, size=9, italic=True)
    scope_c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[3].height = 55

    # KPI block ─────────────────────────────────────────────────────────────
    def _kpi_block(row, col, label, value, bg, fg=C_WHITE):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row+1, end_column=col)
        vc = ws.cell(row=row, column=col, value=value)
        vc.font      = Font(name=FONT_NAME, bold=True, size=22, color=fg)
        vc.fill      = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")

        lc = ws.cell(row=row+2, column=col, value=label)
        lc.font      = Font(name=FONT_NAME, size=9, bold=True, color=bg)
        lc.alignment = Alignment(horizontal="center")

    KPI_ROW = 5
    _kpi_block(KPI_ROW, 1, "Total Requirements", total,       C_SUMMARY_HEAD)
    _kpi_block(KPI_ROW, 2, "Gaps Identified",    n_gap,       C_GAP_HEAD)
    _kpi_block(KPI_ROW, 3, "Standard SAP (FULL)",n_full,      C_NOGAP_HEAD)
    _kpi_block(KPI_ROW, 4, "Partial Gaps",        n_partial,  "E36209")
    _kpi_block(KPI_ROW, 5, "Custom Dev Required", n_none,     "7B0000")
    _kpi_block(KPI_ROW, 6, "Open Actions",        len(actions),C_ACTION_HEAD)

    for r in range(KPI_ROW, KPI_ROW + 3):
        ws.row_dimensions[r].height = 24

    # RICEFW breakdown ─────────────────────────────────────────────────────
    row = KPI_ROW + 4
    ws.cell(row=row, column=1, value="RICEFW Breakdown").font = Font(
        name=FONT_NAME, bold=True, size=11, color=C_SUMMARY_HEAD)
    ws.row_dimensions[row].height = 20
    row += 1

    ricefw_colors = {
        "Report":      "9DC3E6", "Interface":   "F4B942",
        "Conversion":  "A9D18E", "Enhancement": "F4B8A8",
        "Form":        "FFE699", "Workflow":    "D5A6BD",
    }
    for col_idx, (rf, cnt) in enumerate(sorted(ricefw_counts.items()), start=1):
        hc = ws.cell(row=row,   column=col_idx, value=rf)
        vc = ws.cell(row=row+1, column=col_idx, value=cnt)
        bg = ricefw_colors.get(rf, "E0E0E0")
        hdr(hc, bg=C_NAVY, size=9)
        fill(vc, bg)
        vc.font      = Font(name=FONT_NAME, bold=True, size=14)
        vc.alignment = Alignment(horizontal="center")

    ws.row_dimensions[row].height   = 18
    ws.row_dimensions[row+1].height = 24

    # Gap list quick-ref ───────────────────────────────────────────────────
    row = row + 3
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="Gap Quick Reference").font = Font(
        name=FONT_NAME, bold=True, size=11, color=C_SUMMARY_HEAD)
    ws.row_dimensions[row].height = 20
    row += 1

    for col_idx, label in enumerate(
        ["GAP ID", "Req ID", "Title", "RICEFW", "SAP Status", "Sheet Ref"], start=1
    ):
        c = ws.cell(row=row, column=col_idx, value=label)
        hdr(c, bg=C_GAP_HEAD, size=FONT_SIZE)
    ws.row_dimensions[row].height = 20
    row += 1

    cap_map = {c["id"]: c for c in caps}
    for g in data.get("gap_analysis", []):
        cap   = cap_map.get(g.get("req_id", ""), {})
        bg    = "F4B8A8" if cap.get("status") == "NONE" else C_GAP_ROW
        cells = [g.get("gap_id",""), g.get("req_id",""), g.get("title",""),
                 g.get("ricefw",""), cap.get("status",""), "Gap Analysis (RICEFW)"]
        for col_idx, val in enumerate(cells, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            body(c)
            fill(c, bg)
        row += 1

    apply_border(ws, 4, ws.max_row, 1, 6)

    col_widths(ws, {
        "A": 12, "B": 10, "C": 40, "D": 16, "E": 14, "F": 24
    })


# ── Main builder ─────────────────────────────────────────────────────────────
def build_excel(json_path: str, output_path: str) -> str:
    data = json.loads(Path(json_path).read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    _sheet_summary(wb, data)
    _sheet_master(wb, data)
    _sheet_gaps(wb, data)
    _sheet_nogap(wb, data)
    _sheet_actions(wb, data)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"[OK] Written → {output_path}")
    return output_path


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert sap_gap_analyser step_final.json to Excel"
    )
    parser.add_argument("--json",   required=True, help="Path to step_final.json")
    parser.add_argument("--output", default="SAP_Gap_Analysis.xlsx",
                        help="Output .xlsx path")
    args = parser.parse_args()
    build_excel(args.json, args.output)
